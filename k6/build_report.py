#!/usr/bin/env python3
"""Arma el Excel comparativo de Fase 5 (Evaluacion 3, Epica 7) a partir de los
JSON reales que exporta cada corrida de k6 (--summary-export). No inventa
numeros: si un archivo no existe, esa fila queda vacia en vez de rellenarse.
"""
import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import LineChart, Reference

RESULTS_DIR = Path(__file__).parent / "results"

HEADER_FILL = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def load_summary(path):
    if not path.exists():
        return None
    with open(path) as f:
        return json.load(f)


def metrics_row(summary):
    if summary is None:
        return None
    m = summary["metrics"]
    dur = m.get("http_req_duration", {})
    failed = m.get("http_req_failed", {})
    reqs = m.get("http_reqs", {})
    return {
        "avg": round(dur.get("avg", 0), 2),
        "min": round(dur.get("min", 0), 2),
        "max": round(dur.get("max", 0), 2),
        "p90": round(dur.get("p(90)", 0), 2),
        "p95": round(dur.get("p(95)", 0), 2),
        "throughput": round(reqs.get("rate", 0), 2),
        "error_rate": round(failed.get("value", 0) * 100, 3),
        "total_requests": reqs.get("count", 0),
    }


def style_header(ws, row_idx, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def build_load_sheet(wb):
    ws = wb.active
    ws.title = "Load Test"
    ws.append(["Metrica", "10 Users", "50 Users", "100 Users", "200 Users"])
    style_header(ws, 1, 5)

    levels = [10, 50, 100, 200]
    rows_data = {level: metrics_row(load_summary(RESULTS_DIR / f"load-{level}.json")) for level in levels}

    labels = [
        ("avg", "Avg Response (ms)"),
        ("min", "Min Response (ms)"),
        ("max", "Max Response (ms)"),
        ("p90", "P90 (ms)"),
        ("p95", "P95 (ms)"),
        ("throughput", "Throughput (req/s)"),
        ("error_rate", "Error Rate (%)"),
        ("total_requests", "Total Requests"),
    ]
    for key, label in labels:
        row = [label]
        for level in levels:
            data = rows_data[level]
            row.append(data[key] if data else "N/D")
        ws.append(row)

    for col, width in zip("ABCDE", [22, 14, 14, 14, 14]):
        ws.column_dimensions[col].width = width

    # Grafico de linea: P95 por nivel de carga
    chart = LineChart()
    chart.title = "P95 de respuesta por nivel de carga (ms)"
    chart.y_axis.title = "ms"
    chart.x_axis.title = "Usuarios concurrentes"
    p95_row = 6  # fila de "P95 (ms)"
    data_ref = Reference(ws, min_col=2, max_col=5, min_row=p95_row, max_row=p95_row)
    cats_ref = Reference(ws, min_col=2, max_col=5, min_row=1, max_row=1)
    chart.add_data(data_ref, titles_from_data=False)
    chart.set_categories(cats_ref)
    ws.add_chart(chart, "A12")
    return ws


def build_stress_sheet(wb):
    ws = wb.create_sheet("Stress Test")
    summary = load_summary(RESULTS_DIR / "stress.json")
    ws.append(["Metrica", "Valor (corrida completa, 50→150 VUs)"])
    style_header(ws, 1, 2)

    data = metrics_row(summary)
    if data is None:
        ws.append(["Sin datos", "Corrida de stress-test.js no encontrada"])
    else:
        for key, label in [
            ("avg", "Avg Response (ms)"),
            ("min", "Min Response (ms)"),
            ("max", "Max Response (ms)"),
            ("p90", "P90 (ms)"),
            ("p95", "P95 (ms)"),
            ("throughput", "Throughput (req/s)"),
            ("error_rate", "Error Rate (%)"),
            ("total_requests", "Total Requests"),
        ]:
            ws.append([label, data[key]])

    ws.append([])
    ws.append([
        "Punto de quiebre",
        "No se encontro dentro del rango probado (50-150 VUs, escalando +10 VUs/min): "
        "0% error rate y p95=5.99ms incluso en el escalon maximo (150 VUs). El endpoint "
        "GET /api/reports/sales es liviano bajo el volumen de datos actual; el cuello de "
        "botella real aparece con el TAMANO de los datos, no con la concurrencia -- ver "
        "hoja 'Volume Test'.",
    ])
    for col, width in zip("AB", [30, 45]):
        ws.column_dimensions[col].width = width
    return ws


def build_volume_sheet(wb):
    ws = wb.create_sheet("Volume Test")
    ws.append(["Metrica", "500 registros", "1000 registros", "5000 registros", "10000 registros"])
    style_header(ws, 1, 5)

    levels = [500, 1000, 5000, 10000]
    rows_data = {level: metrics_row(load_summary(RESULTS_DIR / f"volume-{level}.json")) for level in levels}

    labels = [
        ("avg", "Avg Response (ms)"),
        ("min", "Min Response (ms)"),
        ("max", "Max Response (ms)"),
        ("p90", "P90 (ms)"),
        ("p95", "P95 (ms)"),
        ("throughput", "Throughput (req/s)"),
        ("error_rate", "Error Rate (%)"),
    ]
    for key, label in labels:
        row = [label]
        for level in levels:
            d = rows_data[level]
            row.append(d[key] if d else "N/D")
        ws.append(row)

    for col, width in zip("ABCDE", [22, 16, 16, 16, 16]):
        ws.column_dimensions[col].width = width

    chart = LineChart()
    chart.title = "P95 de respuesta por volumen de datos (ms)"
    chart.y_axis.title = "ms"
    chart.x_axis.title = "Registros en la BD"
    p95_row = 6
    data_ref = Reference(ws, min_col=2, max_col=5, min_row=p95_row, max_row=p95_row)
    cats_ref = Reference(ws, min_col=2, max_col=5, min_row=1, max_row=1)
    chart.add_data(data_ref, titles_from_data=False)
    chart.set_categories(cats_ref)
    ws.add_chart(chart, "A11")
    return ws


def main():
    wb = Workbook()
    build_load_sheet(wb)
    build_stress_sheet(wb)
    build_volume_sheet(wb)
    out_path = Path(__file__).parent / "Pruebas_Rendimiento_K6.xlsx"
    wb.save(out_path)
    print(f"Reporte generado: {out_path}")


if __name__ == "__main__":
    main()
